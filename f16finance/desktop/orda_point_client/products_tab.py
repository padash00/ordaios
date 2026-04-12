"""
Products management tab for Orda Control Point
"""
from __future__ import annotations

import uuid
from typing import Optional, List, Dict, Any, Tuple
from enum import Enum
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

import theme
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QSortFilterProxyModel
from PyQt6.QtGui import QIntValidator, QColor, QPalette, QFont, QIcon, QPixmap
from utils import parse_money, format_money
from PyQt6.QtWidgets import (
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QCheckBox,
    QFrame,
    QGroupBox,
    QSplitter,
    QProgressBar,
    QComboBox,
    QApplication,
)


# ==================== Вспомогательные функции ====================

def generate_barcode() -> str:
    """Генерация случайного штрихкода"""
    import random
    return ''.join([str(random.randint(0, 9)) for _ in range(13)])


class ProductStatus(Enum):
    """Статус товара"""
    ACTIVE = "active"
    INACTIVE = "inactive"
    
    def display_name(self) -> str:
        names = {
            "active": "Активен",
            "inactive": "Выключен"
        }
        return names.get(self.value, self.value)
    
    def color(self) -> str:
        colors = {
            "active": "#10B981",
            "inactive": "#EF4444"
        }
        return colors.get(self.value, "#93A5C1")


# ==================== Современные UI компоненты ====================

class StatsCard(QFrame):
    """Карточка статистики товаров"""
    
    def __init__(self, title: str, icon: str, parent=None):
        super().__init__(parent)
        self.setProperty("class", "stats-card")

        self.setStyleSheet(f"""
            StatsCard {{
                background: {theme.CARD};
                border: 1px solid {theme.DIVIDER};
                border-radius: 16px;
                padding: 16px;
            }}
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(12)
        
        # Иконка
        icon_label = QLabel(icon)
        icon_label.setStyleSheet("font-size: 28px; background: transparent;")
        
        # Текстовый контейнер
        text_container = QVBoxLayout()
        text_container.setSpacing(4)
        
        self.title_label = QLabel(title)
        self.title_label.setProperty("class", "muted")
        self.title_label.setStyleSheet("font-size: 12px; background: transparent;")
        
        self.value_label = QLabel("0")
        self.value_label.setStyleSheet(f"""
            font-size: 24px;
            font-weight: 700;
            color: {theme.ACCENT};
            background: transparent;
        """)
        
        text_container.addWidget(self.title_label)
        text_container.addWidget(self.value_label)
        
        layout.addWidget(icon_label)
        layout.addLayout(text_container, 1)
    
    def set_value(self, value: int | str):
        """Установка значения"""
        self.value_label.setText(str(value))


class ProductForm(QGroupBox):
    """
    Улучшенная форма для создания/редактирования товаров
    
    Сигналы:
        product_saved: испускается при сохранении товара
        form_changed: испускается при изменении формы
    """
    
    product_saved = pyqtSignal(dict)
    form_changed = pyqtSignal()
    
    def __init__(self, parent=None):
        super().__init__("Добавить / Редактировать товар", parent)
        self.editing_id: Optional[str] = None
        self.setup_ui()

    def setup_ui(self):
        """Настройка интерфейса формы"""
        self.setStyleSheet(f"""
            ProductForm {{
                border: 1px solid {theme.DIVIDER};
                border-radius: 16px;
                margin-top: 16px;
                font-weight: 600;
                color: {theme.ACCENT};
                background: {theme.CARD};
                padding: 20px;
            }}
            ProductForm::title {{
                subcontrol-origin: margin;
                left: 16px;
                padding: 0 12px;
                background: {theme.BG};
                font-size: 15px;
            }}
        """)
        
        layout = QFormLayout(self)
        layout.setVerticalSpacing(16)
        layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        
        # Название товара
        name_label = QLabel("Название")
        name_label.setStyleSheet(f"font-size: 14px; font-weight: 600; color: {theme.ACCENT};")

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Введите название товара")
        self.name_input.setMinimumHeight(44)
        self.name_input.textChanged.connect(self.on_form_changed)

        # Штрихкод
        barcode_label = QLabel("Штрихкод")
        barcode_label.setStyleSheet(f"font-size: 14px; font-weight: 600; color: {theme.ACCENT};")
        
        barcode_container = QWidget()
        barcode_layout = QHBoxLayout(barcode_container)
        barcode_layout.setContentsMargins(0, 0, 0, 0)
        barcode_layout.setSpacing(8)
        
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("Введите или сгенерируйте штрихкод")
        self.barcode_input.setMinimumHeight(44)
        self.barcode_input.textChanged.connect(self.on_form_changed)

        self.generate_barcode_btn = QPushButton("Генерировать")
        self.generate_barcode_btn.setFixedHeight(44)
        self.generate_barcode_btn.setToolTip("Сгенерировать случайный штрихкод")
        self.generate_barcode_btn.setStyleSheet(f"""
            QPushButton {{
                background: {theme.CARD};
                border: 1px solid {theme.DIVIDER};
                border-radius: 10px;
                font-size: 13px;
                padding: 0 12px;
                color: {theme.TEXT};
            }}
            QPushButton:hover {{
                background: {theme.DIVIDER};
            }}
        """)
        self.generate_barcode_btn.clicked.connect(self.generate_barcode)
        
        barcode_layout.addWidget(self.barcode_input, 1)
        barcode_layout.addWidget(self.generate_barcode_btn)
        
        # Цена
        price_label = QLabel("Цена")
        price_label.setStyleSheet(f"font-size: 14px; font-weight: 600; color: {theme.ACCENT};")

        price_container = QWidget()
        price_layout = QHBoxLayout(price_container)
        price_layout.setContentsMargins(0, 0, 0, 0)
        price_layout.setSpacing(8)

        self.price_input = QLineEdit("0")
        self.price_input.setValidator(QIntValidator(0, 9_999_999))
        self.price_input.setMinimumHeight(44)
        self.price_input.textChanged.connect(self.on_form_changed)

        currency_label = QLabel("₸")
        currency_label.setStyleSheet(f"""
            font-size: 18px;
            font-weight: 700;
            color: {theme.WARNING};
            background: transparent;
        """)

        price_layout.addWidget(self.price_input, 1)
        price_layout.addWidget(currency_label)

        # Статус
        status_label = QLabel("Статус")
        status_label.setStyleSheet(f"font-size: 14px; font-weight: 600; color: {theme.ACCENT};")
        
        status_container = QWidget()
        status_layout = QHBoxLayout(status_container)
        status_layout.setContentsMargins(0, 0, 0, 0)
        status_layout.setSpacing(12)
        
        self.active_check = QCheckBox("Активный товар")
        self.active_check.setChecked(True)
        self.active_check.setStyleSheet(f"""
            QCheckBox {{
                color: {theme.SUCCESS};
                font-size: 14px;
                spacing: 8px;
            }}
            QCheckBox::indicator {{
                width: 20px;
                height: 20px;
                border: 1px solid {theme.DIVIDER};
                border-radius: 5px;
                background: {theme.CARD};
            }}
            QCheckBox::indicator:checked {{
                background: {theme.SUCCESS};
                border-color: {theme.SUCCESS};
            }}
            QCheckBox::indicator:hover {{
                border-color: {theme.ACCENT};
            }}
        """)
        self.active_check.stateChanged.connect(self.on_form_changed)
        
        self.status_badge = QFrame()
        self.status_badge.setFixedSize(8, 8)
        self.status_badge.setStyleSheet("background: #10B981; border-radius: 4px;")
        
        status_layout.addWidget(self.active_check)
        status_layout.addWidget(self.status_badge)
        status_layout.addStretch()
        
        # Добавление полей в форму
        layout.addRow(name_label, self.name_input)
        layout.addRow(barcode_label, barcode_container)
        layout.addRow(price_label, price_container)
        layout.addRow(status_label, status_container)
        
        # Индикатор изменений
        self.changed_indicator = QLabel("Есть несохранённые изменения")
        self.changed_indicator.setStyleSheet(f"""
            font-size: 12px;
            color: {theme.WARNING};
            background: transparent;
            padding: 4px 8px;
        """)
        self.changed_indicator.hide()
        layout.addRow("", self.changed_indicator)

        # Подсказка
        hint_label = QLabel("Для массового добавления используйте импорт из Excel")
        hint_label.setProperty("class", "muted")
        hint_label.setStyleSheet(f"""
            font-size: 12px;
            color: {theme.TEXT_MUTED};
            background: transparent;
            padding: 8px 0;
        """)
        layout.addRow("", hint_label)
        
    def on_form_changed(self):
        """Обработка изменения формы"""
        has_content = bool(self.name_input.text().strip() or 
                          self.barcode_input.text().strip() or 
                          self.price_input.text() != "0")
        self.changed_indicator.setVisible(has_content)
        self.form_changed.emit()
        
    def generate_barcode(self):
        """Генерация случайного штрихкода"""
        self.barcode_input.setText(generate_barcode())
        
    def load_product(self, product: dict):
        """Загрузка товара в форму"""
        self.editing_id = str(product.get("id") or "")
        self.name_input.setText(str(product.get("name") or ""))
        self.barcode_input.setText(str(product.get("barcode") or ""))
        self.price_input.setText(str(int(product.get("price") or 0)))
        
        is_active = product.get("is_active") is not False
        self.active_check.setChecked(is_active)
        self.update_status_badge(is_active)
        
    def update_status_badge(self, is_active: bool):
        """Обновление бейджа статуса"""
        color = "#10B981" if is_active else "#EF4444"
        self.status_badge.setStyleSheet(f"background: {color}; border-radius: 4px;")
        
    def get_payload(self) -> Optional[Dict]:
        """Получение данных из формы"""
        name = self.name_input.text().strip()
        barcode = self.barcode_input.text().strip()
        
        try:
            price = parse_money(self.price_input.text())
        except ValueError:
            price = 0
            
        if not name:
            QMessageBox.warning(self, "Товары", "Введите название товара.")
            return None
            
        if not barcode:
            QMessageBox.warning(self, "Товары", "Введите штрихкод.")
            return None
            
        if price <= 0:
            QMessageBox.warning(self, "Товары", "Цена должна быть больше нуля.")
            return None
            
        return {
            "name": name,
            "barcode": barcode,
            "price": price,
            "is_active": self.active_check.isChecked(),
        }
        
    def clear(self):
        """Очистка формы"""
        self.editing_id = None
        self.name_input.clear()
        self.barcode_input.clear()
        self.price_input.setText("0")
        self.active_check.setChecked(True)
        self.update_status_badge(True)
        self.changed_indicator.hide()
        
    def is_editing(self) -> bool:
        """Проверка, редактируется ли существующий товар"""
        return self.editing_id is not None


class ProductsTable(QTableWidget):
    """
    Улучшенная таблица товаров
    
    Сигналы:
        product_selected: испускается при выборе товара
        product_double_clicked: испускается при двойном клике
    """
    
    product_selected = pyqtSignal(dict)
    product_double_clicked = pyqtSignal(dict)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.products: List[Dict] = []
        self.setup_style()
        
    def setup_style(self):
        """Настройка стиля таблицы"""
        self.setShowGrid(False)
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.setVerticalScrollMode(QTableWidget.ScrollMode.ScrollPerPixel)
        
        self.setStyleSheet(f"""
            QTableWidget {{
                background: {theme.BG};
                border: 1px solid {theme.DIVIDER};
                border-radius: 14px;
                gridline-color: transparent;
                selection-background-color: rgba(124,58,237,0.15);
            }}
            QTableWidget::item {{
                padding: 12px 8px;
                border-bottom: 1px solid {theme.DIVIDER};
            }}
            QTableWidget::item:selected {{
                background: rgba(124,58,237,0.15);
            }}
            QTableWidget::item:hover {{
                background: {theme.CARD};
            }}
            QTableWidget::item:selected:hover {{
                background: rgba(124,58,237,0.2);
            }}
        """)

        # Настройка заголовков
        header = self.horizontalHeader()
        header.setStyleSheet(f"""
            QHeaderView::section {{
                background: {theme.BG};
                color: {theme.TEXT_MUTED};
                border: none;
                border-bottom: 2px solid {theme.DIVIDER};
                padding: 14px 8px;
                font-weight: 700;
                font-size: 13px;
            }}
        """)

        # Подключение сигналов
        self.itemSelectionChanged.connect(self.on_selection_changed)
        self.cellDoubleClicked.connect(self.on_double_click)
        
    def set_products(self, products: List[Dict]):
        """Установка списка товаров"""
        self.products = products
        self.update_table()
        
    def update_table(self):
        """Обновление таблицы"""
        self.setRowCount(len(self.products))
        self.setColumnCount(5)
        self.setHorizontalHeaderLabels(["ID", "Название", "Штрихкод", "Цена", "Статус"])
        self.setColumnHidden(0, True)  # Скрываем ID
        
        for row, product in enumerate(self.products):
            # ID (скрыт)
            self.setItem(row, 0, QTableWidgetItem(str(product.get("id") or "")))
            
            # Название
            name_item = QTableWidgetItem(product.get('name', ''))
            self.setItem(row, 1, name_item)

            # Штрихкод
            barcode = product.get('barcode', '')
            barcode_item = QTableWidgetItem(barcode)
            
            # Визуальное выделение для штрихкодов
            if barcode and len(barcode) == 13:
                barcode_item.setForeground(QColor("#3B82F6"))
                
            self.setItem(row, 2, barcode_item)
            
            # Цена
            price = int(product.get('price') or 0)
            price_item = QTableWidgetItem(f"{format_money(price)} ₸")
            price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            
            # Цветовая индикация цен
            if price > 100000:
                price_item.setForeground(QColor("#F59E0B"))
            elif price > 50000:
                price_item.setForeground(QColor("#3B82F6"))
                
            self.setItem(row, 3, price_item)
            
            # Статус с бейджем
            is_active = product.get('is_active') is not False
            status = "Активен" if is_active else "Выключен"
            status_item = QTableWidgetItem(status)
            status_item.setForeground(QColor("#10B981" if is_active else "#EF4444"))
            self.setItem(row, 4, status_item)
            
        # Настройка растяжения колонок
        header = self.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Название
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Штрихкод
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Цена
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Статус
        
    def on_selection_changed(self):
        """Обработка изменения выбора"""
        row = self.currentRow()
        if 0 <= row < len(self.products):
            self.product_selected.emit(self.products[row])
            
    def on_double_click(self, row: int, column: int):
        """Обработка двойного клика"""
        if 0 <= row < len(self.products):
            self.product_double_clicked.emit(self.products[row])
            
    def selected_product(self) -> Optional[Dict]:
        """Получение выбранного товара"""
        row = self.currentRow()
        if 0 <= row < len(self.products):
            return self.products[row]
        return None


class ImportProgressDialog(QMessageBox):
    """Диалог с прогрессом импорта"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Импорт товаров")
        self.setIcon(QMessageBox.Icon.Information)
        
        # Кастомный контент
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(8)
        self.progress_bar.setStyleSheet(f"""
            QProgressBar {{
                background: {theme.CARD};
                border: none;
                border-radius: 4px;
                text-align: center;
            }}
            QProgressBar::chunk {{
                background: {theme.ACCENT};
                border-radius: 4px;
            }}
        """)
        
        # Добавляем в диалог
        layout = self.layout()
        if layout:
            layout.addRow(self.progress_bar)
            
    def set_progress(self, value: int):
        """Установка прогресса"""
        self.progress_bar.setValue(value)
        QApplication.processEvents()


# ==================== Основной класс вкладки ====================

class ProductsTab(QWidget):
    """
    Улучшенная вкладка управления товарами
    
    Особенности:
    - Современный дизайн
    - Статистика товаров
    - Улучшенная форма
    - Поиск и фильтрация
    - Прогресс импорта
    - Экспорт в Excel
    """
    
    def __init__(self, main_window):
        super().__init__(main_window)
        self.main_window = main_window
        self.products: List[Dict] = []
        self.init_ui()
        self.load_products()
        
    def init_ui(self):
        """Инициализация интерфейса"""
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 20)
        root.setSpacing(20)

        # === Верхняя панель статистики ===
        stats_layout = QHBoxLayout()
        stats_layout.setSpacing(16)
        
        self.total_card = StatsCard("Всего товаров", "📊")
        self.active_card = StatsCard("Активных", "✅")
        self.inactive_card = StatsCard("Неактивных", "⭕")
        self.total_value_card = StatsCard("Общая стоимость", "💰")
        
        stats_layout.addWidget(self.total_card)
        stats_layout.addWidget(self.active_card)
        stats_layout.addWidget(self.inactive_card)
        stats_layout.addWidget(self.total_value_card)
        
        root.addLayout(stats_layout)

        # === Панель поиска ===
        search_container = QFrame()
        search_container.setStyleSheet(f"""
            QFrame {{
                background: {theme.BG};
                border: 1px solid {theme.DIVIDER};
                border-radius: 12px;
                padding: 8px 12px;
            }}
        """)
        
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(12, 8, 12, 8)
        
        search_icon = QLabel(">>")
        search_icon.setStyleSheet(f"font-size: 13px; color: {theme.TEXT_MUTED}; background: transparent;")

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск по названию или штрихкоду...")
        self.search_input.setMinimumHeight(36)
        self.search_input.textChanged.connect(self.filter_products)

        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["Все товары", "Только активные", "Только неактивные"])
        self.filter_combo.setMinimumHeight(36)
        self.filter_combo.setMinimumWidth(150)
        self.filter_combo.currentTextChanged.connect(self.filter_products)
        
        search_layout.addWidget(search_icon)
        search_layout.addWidget(self.search_input, 1)
        search_layout.addWidget(self.filter_combo)
        
        root.addWidget(search_container)

        # === Форма товара ===
        self.product_form = ProductForm()
        self.product_form.form_changed.connect(self.update_save_button)
        root.addWidget(self.product_form)

        # === Панель действий ===
        actions_layout = QHBoxLayout()
        actions_layout.setSpacing(12)
        
        # Левая группа
        left_actions = QHBoxLayout()
        left_actions.setSpacing(8)
        
        self.save_btn = QPushButton("Сохранить")
        self.save_btn.setProperty("class", "success")
        self.save_btn.setMinimumHeight(44)
        self.save_btn.setMinimumWidth(140)
        self.save_btn.clicked.connect(self.save_product)

        self.clear_btn = QPushButton("Очистить")
        self.clear_btn.setProperty("class", "ghost")
        self.clear_btn.setMinimumHeight(44)
        self.clear_btn.clicked.connect(self.clear_form)
        
        left_actions.addWidget(self.save_btn)
        left_actions.addWidget(self.clear_btn)
        
        # Центр
        center_actions = QHBoxLayout()
        center_actions.setSpacing(8)
        
        self.reload_btn = QPushButton("Обновить")
        self.reload_btn.setProperty("class", "ghost")
        self.reload_btn.setMinimumHeight(44)
        self.reload_btn.clicked.connect(self.load_products)
        
        center_actions.addWidget(self.reload_btn)
        
        # Правая группа
        right_actions = QHBoxLayout()
        right_actions.setSpacing(8)
        
        self.import_btn = QPushButton("Импорт Excel")
        self.import_btn.setProperty("class", "primary")
        self.import_btn.setMinimumHeight(44)
        self.import_btn.clicked.connect(self.import_from_excel)

        self.export_btn = QPushButton("Экспорт Excel")
        self.export_btn.setProperty("class", "ghost")
        self.export_btn.setMinimumHeight(44)
        self.export_btn.clicked.connect(self.export_to_excel)

        self.delete_btn = QPushButton("Удалить")
        self.delete_btn.setProperty("class", "danger")
        self.delete_btn.setMinimumHeight(44)
        self.delete_btn.clicked.connect(self.delete_selected)
        
        right_actions.addWidget(self.import_btn)
        right_actions.addWidget(self.export_btn)
        right_actions.addWidget(self.delete_btn)
        
        actions_layout.addLayout(left_actions)
        actions_layout.addStretch(1)
        actions_layout.addLayout(center_actions)
        actions_layout.addStretch(1)
        actions_layout.addLayout(right_actions)
        
        root.addLayout(actions_layout)

        # === Информационная строка ===
        info_container = QFrame()
        info_container.setStyleSheet(f"""
            QFrame {{
                background: rgba(124, 58, 237, 0.05);
                border: 1px solid rgba(124, 58, 237, 0.2);
                border-radius: 10px;
                padding: 8px 12px;
            }}
        """)

        info_layout = QHBoxLayout(info_container)
        info_layout.setContentsMargins(12, 8, 12, 8)

        info_icon = QLabel("i")
        info_icon.setStyleSheet(f"font-size: 13px; color: {theme.ACCENT}; font-weight: 700;")

        self.info_label = QLabel("Каталог товаров текущей точки")
        self.info_label.setProperty("class", "muted")
        self.info_label.setStyleSheet("font-size: 13px;")
        
        info_layout.addWidget(info_icon)
        info_layout.addWidget(self.info_label, 1)
        
        root.addWidget(info_container)

        # === Таблица товаров ===
        self.table = ProductsTable()
        self.table.product_selected.connect(self.on_product_selected)
        self.table.product_double_clicked.connect(self.on_product_double_clicked)
        root.addWidget(self.table, 1)

        # === Статусная строка внизу ===
        status_container = QFrame()
        status_container.setStyleSheet(f"""
            QFrame {{
                background: {theme.BG};
                border: 1px solid {theme.DIVIDER};
                border-radius: 8px;
                padding: 6px 12px;
            }}
        """)
        
        status_layout = QHBoxLayout(status_container)
        status_layout.setContentsMargins(12, 6, 12, 6)
        
        self.status_label = QLabel("Готов к работе")
        self.status_label.setProperty("class", "muted")
        self.status_label.setStyleSheet("font-size: 12px;")
        
        self.last_update_label = QLabel("")
        self.last_update_label.setProperty("class", "muted")
        self.last_update_label.setStyleSheet("font-size: 12px;")
        
        status_layout.addWidget(self.status_label)
        status_layout.addStretch(1)
        status_layout.addWidget(self.last_update_label)
        
        root.addWidget(status_container)

    # ==================== Методы работы с данными ====================

    def require_admin(self):
        """Проверка прав администратора"""
        creds = self.main_window.admin_credentials
        if not creds or not self.main_window.current_admin:
            QMessageBox.warning(self, "Товары", "Сначала войдите как super-admin.")
            return None
        return creds

    def load_products(self):
        """Загрузка списка товаров"""
        if not self.main_window.api:
            self.products = []
            self.update_table()
            return

        try:
            response = self.main_window.api.list_products()
            self.products = ((response.get("data") or {}).get("products") or [])
            
            # Обновление статистики
            self.update_statistics()
            
            self.info_label.setText(f"📦 Товаров в каталоге: {len(self.products)}")
            self.status_label.setText("✓ Данные загружены")
            self.last_update_label.setText(f"Обновлено: {datetime.now().strftime('%H:%M:%S')}")
            
        except Exception as error:
            self.products = []
            self.info_label.setText(f"⚠️ Каталог недоступен: {error}")
            self.status_label.setText("✗ Ошибка загрузки")

        self.update_table()
        
    def update_table(self):
        """Обновление таблицы"""
        self.table.set_products(self.products)
        self.apply_filter()  # Применяем текущий фильтр
        
    def update_statistics(self):
        """Обновление статистики"""
        total = len(self.products)
        active = sum(1 for p in self.products if p.get("is_active") is not False)
        inactive = total - active
        total_value = sum(int(p.get("price") or 0) for p in self.products)
        
        self.total_card.set_value(total)
        self.active_card.set_value(active)
        self.inactive_card.set_value(inactive)
        self.total_value_card.set_value(f"{format_money(total_value)} ₸")
        
    def filter_products(self):
        """Фильтрация товаров"""
        search_text = self.search_input.text().lower()
        filter_type = self.filter_combo.currentText()
        
        filtered = []
        for product in self.products:
            # Фильтр по статусу
            if filter_type == "Только активные" and product.get("is_active") is False:
                continue
            if filter_type == "Только неактивные" and product.get("is_active") is not False:
                continue
                
            # Поиск по тексту
            if search_text:
                name = str(product.get("name", "")).lower()
                barcode = str(product.get("barcode", "")).lower()
                if search_text not in name and search_text not in barcode:
                    continue
                    
            filtered.append(product)
            
        self.table.set_products(filtered)
        self.info_label.setText(f"📦 Показано: {len(filtered)} из {len(self.products)}")
        
    def update_save_button(self):
        """Обновление состояния кнопки сохранения"""
        has_content = bool(self.product_form.name_input.text().strip())
        self.save_btn.setEnabled(has_content)

    # ==================== Обработчики событий ====================

    def on_product_selected(self, product: dict):
        """Обработка выбора товара"""
        self.product_form.load_product(product)
        self.update_save_button()
        
    def on_product_double_clicked(self, product: dict):
        """Обработка двойного клика по товару"""
        self.product_form.load_product(product)
        self.update_save_button()
        # Можно добавить фокус на форму
        self.product_form.name_input.setFocus()

    def save_product(self):
        """Сохранение товара"""
        creds = self.require_admin()
        if not creds or not self.main_window.api:
            return

        payload = self.product_form.get_payload()
        if not payload:
            return

        try:
            if self.product_form.is_editing():
                # Обновление существующего
                self.main_window.api.update_product(
                    creds["email"],
                    creds["password"],
                    self.product_form.editing_id,
                    payload,
                )
                self.status_label.setText("✓ Товар обновлён")
            else:
                # Создание нового
                self.main_window.api.create_product(
                    creds["email"],
                    creds["password"],
                    payload,
                )
                self.status_label.setText("✓ Товар добавлен")
                
            # Обновление данных
            self.load_products()
            if self.main_window.scanner_tab:
                self.main_window.scanner_tab.load_products()
                
            self.clear_form()
            
            # Показываем успех
            QMessageBox.information(
                self, 
                "Товары", 
                "Товар успешно сохранён в каталоге."
            )
            
        except Exception as error:
            self.status_label.setText(f"✗ Ошибка: {error}")
            QMessageBox.critical(self, "Товары", str(error))

    def clear_form(self):
        """Очистка формы"""
        self.product_form.clear()
        self.table.clearSelection()
        self.update_save_button()

    def delete_selected(self):
        """Удаление выбранного товара"""
        creds = self.require_admin()
        product = self.table.selected_product()
        
        if not creds or not product or not self.main_window.api:
            return

        # Подтверждение
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Удалить товар «{product.get('name')}»?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            self.main_window.api.delete_product(
                creds["email"],
                creds["password"],
                str(product.get("id") or ""),
            )
            
            self.load_products()
            if self.main_window.scanner_tab:
                self.main_window.scanner_tab.load_products()
                
            self.clear_form()
            self.status_label.setText("✓ Товар удалён")
            
            QMessageBox.information(self, "Товары", "Товар удалён.")
            
        except Exception as error:
            self.status_label.setText(f"✗ Ошибка: {error}")
            QMessageBox.critical(self, "Товары", str(error))

    def import_from_excel(self):
        """Импорт товаров из Excel"""
        if not _OPENPYXL_AVAILABLE:
            QMessageBox.critical(
                self, 
                "Excel импорт",
                "Библиотека openpyxl не установлена.\n\nУстановите: pip install openpyxl"
            )
            return

        creds = self.require_admin()
        if not creds or not self.main_window.api:
            return

        # Выбор файла
        path, _ = QFileDialog.getOpenFileName(
            self, 
            "Выберите Excel файл", 
            "", 
            "Excel файлы (*.xlsx *.xls)"
        )
        
        if not path:
            return

        # Диалог прогресса
        progress = ImportProgressDialog(self)
        progress.setText("Идёт импорт товаров...")
        progress.setStandardButtons(QMessageBox.StandardButton.NoButton)
        progress.show()

        try:
            # Загрузка файла
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            
            # Проверка заголовков
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1, values_only=True))]
            expected = ["Название", "Штрихкод", "Цена"]
            
            if not all(h in headers for h in expected):
                QMessageBox.warning(
                    self,
                    "Неверный формат",
                    f"Ожидаются колонки: {', '.join(expected)}"
                )
                return

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total = len(rows)
            
            added = 0
            skipped = 0
            errors = []

            for idx, row in enumerate(rows):
                # Обновление прогресса
                progress.set_progress(int((idx + 1) / total * 100))
                progress.setText(f"Обработано: {idx + 1} из {total}")
                QApplication.processEvents()

                if not row or all(cell is None for cell in row):
                    continue

                name = str(row[0] or "").strip() if len(row) > 0 else ""
                barcode = str(row[1] or "").strip() if len(row) > 1 else ""
                
                try:
                    price_str = str(row[2] or "0").replace(" ", "").replace(",", ".")
                    price = max(0, int(float(price_str)))
                except (ValueError, TypeError):
                    price = 0

                if not name or not barcode or price <= 0:
                    skipped += 1
                    continue

                try:
                    self.main_window.api.create_product(
                        creds["email"],
                        creds["password"],
                        {
                            "name": name,
                            "barcode": barcode,
                            "price": price,
                            "is_active": True,
                        },
                    )
                    added += 1
                except Exception as e:
                    errors.append(f"Стр.{idx + 2} «{name}»: {e}")

            progress.close()

            # Обновление данных
            self.load_products()
            if self.main_window.scanner_tab:
                self.main_window.scanner_tab.load_products()

            # Отчёт
            summary = f"✅ Добавлено: {added}\n⏭️ Пропущено: {skipped}"
            if errors:
                summary += f"\n\n❌ Ошибок: {len(errors)}\n" + "\n".join(errors[:5])
                
            self.status_label.setText(f"✓ Импорт завершён: +{added} товаров")
            
            QMessageBox.information(self, "Excel импорт завершён", summary)

        except Exception as error:
            progress.close()
            self.status_label.setText("✗ Ошибка импорта")
            QMessageBox.critical(self, "Excel импорт", f"Ошибка: {error}")

    def export_to_excel(self):
        """Экспорт товаров в Excel"""
        if not _OPENPYXL_AVAILABLE:
            QMessageBox.critical(
                self,
                "Excel экспорт",
                "Библиотека openpyxl не установлена.\n\nУстановите: pip install openpyxl"
            )
            return

        if not self.products:
            QMessageBox.warning(self, "Экспорт", "Нет товаров для экспорта.")
            return

        # Выбор файла для сохранения
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить Excel файл",
            f"products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel файлы (*.xlsx)"
        )
        
        if not path:
            return

        try:
            # Создание книги
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Товары"

            # Заголовки
            headers = ["ID", "Название", "Штрихкод", "Цена", "Статус"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Данные
            for row, product in enumerate(self.products, 2):
                ws.cell(row=row, column=1, value=product.get("id", ""))
                ws.cell(row=row, column=2, value=product.get("name", ""))
                ws.cell(row=row, column=3, value=product.get("barcode", ""))
                ws.cell(row=row, column=4, value=product.get("price", 0))
                
                status = "Активен" if product.get("is_active") is not False else "Выключен"
                ws.cell(row=row, column=5, value=status)

            # Автоширина колонок
            for col in range(1, 6):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

            # Сохранение
            wb.save(path)
            
            self.status_label.setText(f"✓ Экспортировано {len(self.products)} товаров")
            
            QMessageBox.information(
                self,
                "Excel экспорт",
                f"✅ Экспортировано {len(self.products)} товаров\n\nСохранено: {path}"
            )

        except Exception as error:
            self.status_label.setText("✗ Ошибка экспорта")
            QMessageBox.critical(self, "Excel экспорт", f"Ошибка: {error}")